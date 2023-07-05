import openpyxl  # 第三方模块,  pip install openpyxl

# 1.创建一个工作簿对象
work_book = openpyxl.Workbook()

# 2.创建表对象
# sheet1 = work_book.create_sheet('表1')
# 如果使用默认的表操作数据, 需要调用工作簿对象的active属性
sheet1 = work_book.active


# 3.操作表中单元格写入数据
# sheet1['A1'] = 'A1'
# sheet1['B7'] = 'B7'

# cell --> 单元格对象, row表示行, column表示列
sheet1.cell(row=1, column=1).value = '111111'
sheet1.cell(row=2, column=2).value = '222222'

data1 = (1, 2, 3, 4, 5)
# data2 = '45678'

# sheet1.append(序列数据)  整行添加数据到表格中去, 括号内部传递序列数据(列表/元祖)
# 通过数据的第一次和第二次数据提取, 会提取到一条一条的数据
sheet1.append(data1)
# sheet1.append(data2)

# 4.保存
work_book.save('实例.xlsx')
