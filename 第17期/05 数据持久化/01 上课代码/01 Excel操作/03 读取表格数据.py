import openpyxl


workbook = openpyxl.load_workbook('实例.xlsx')

print(workbook.sheetnames)

sheet = workbook['Sheet']  # 指定表读取

print(sheet.max_row)  # 最大行
print(sheet.max_column)  # 最大列

# # 读取第一行
# for i in range(1, sheet.max_column + 1):
#     print(sheet.cell(row=1, column=i).value)  # 单元格为空就返回None

# # 读取第一列
# for j in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=j, column=1).value)  # 单元格为空就返回None

for i in range(1, sheet.max_column + 1):
    for j in range(1, sheet.max_row + 1):
        print(sheet.cell(row=i, column=j).value)
